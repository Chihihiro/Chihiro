
from engine import *
df=pd.read_sql("select zhaoyang,standard,nv_source,nv_source_copy2,nv_standard_copy2 from zhaoyang WHERE zhaoyang is not NULL AND version='{}'".format(now),engine5)
df.dropna(axis=0, how='any', thresh=None, subset=None, inplace=True)
df["differ_day_standard"] = (df["standard"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_source"] = (df["nv_source"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_source_copy2"] = (df["nv_source_copy2"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_standard_copy2"] = (df["nv_standard_copy2"] - df["zhaoyang"]).apply(lambda x: x.days)
# df["differ_day_source_haomai"] = (df["haomai"] - df["zhaoyang"])
# df["differ_day_source_jinfuzi"] = (df["jinfuzi"] - df["zhaoyang"])

# df.dropna(axis=0, how='any', thresh=None, subset=None, inplace=True)
# axis 指 轴，0是行，1是列，
# how 是删除条件：any 任意一个为na则删除整行/列,all 整行/列为na才删除
# inplace 是否在原DataFrame 上进行删除，false为否
# df.fillna(df.mean())

yy = df[(df.differ_day_source <100 )& (df.differ_day_source>-100)&(df.differ_day_source_copy2<100)&(df.differ_day_source_copy2>-100)&(df.differ_day_standard_copy2<100)&(df.differ_day_standard_copy2>-100)&(df.differ_day_standard<100)&(df.differ_day_standard>-100)]
zz=yy.iloc[:,[5,6,7,8]]
d=zz[(zz.differ_day_standard!=0)|(zz.differ_day_source_copy2!=0)|(zz.differ_day_source!=0)|(zz.differ_day_standard_copy2!=0)]
standard=to_list(d["differ_day_standard"])
source=to_list(d["differ_day_source"])
source_copy2=to_list(d["differ_day_source_copy2"])
standard_copy2=to_list(d["differ_day_standard_copy2"])
zhaoyang=[0]*len(standard)
y=range(len(standard))

import pylab as pl
x1=sorted(standard)
x2=sorted(source)
x3=sorted(source_copy2)
x4=sorted(standard_copy2)
pl.plot(y,x1,  'c')  # use pylab to plot x and y
pl.plot(y,x2,  'y')
pl.plot(y,x3,  'b')
pl.plot(y,x4,  'm')
pl.plot(y,zhaoyang,'r')
pl.title('zy vs yt')  # give plot a title
pl.xlabel('Red:zhaoyang Corn:standard Yellow:source Blue:source2 Maroon:standard2')  # make axis labels
pl.ylabel('diff day')
pl.xlim(0,len(standard))  # set axis limits
pl.ylim(-100, 100.)
pl.savefig('C:\\Users\\63220\\PycharmProjects\\QQX\\ceshi.png')
pl.show()
import xlrd
import xlsxwriter


from openpyxl import load_workbook
book = xlsxwriter.Workbook("C:\\Users\\63220\\Desktop\\私募净值追踪{}.csv".format(now2))
# sheet = book.add_worksheet()
# sheet = book.add_worksheet('统计图片')
sheet=book.add_worksheet('私募净值追踪{}'.format(now2))
# sheet=book.get_worksheet_by_name('私募净值追踪{}'.format(now2))
sheet.insert_image('Q5','C:\\Users\\63220\\PycharmProjects\\QQX\\ceshi.png' )
book.close()





import xlsxwriter
book = xlsxwriter.Workbook("C:\\Users\\63220\\Desktop\\插入测试.xlsx")
sheet = book.add_worksheet('TEST')
# sheet=book.get_worksheet_by_name('插入')
sheet.insert_image('Q5','C:\\Users\\63220\\PycharmProjects\\QQX\\ceshi.png' )

book.close()



 #coding=utf-8
#/usr/bin/env python
import xlsxwriter
import xlrd
import sys,os.path
fname = "C:\\Users\\63220\\Desktop\\插入测试.xlsx"

if not os.path.isfile(fname):
    print ('文件路径不存在')
    sys.exit()
data = xlrd.open_workbook(fname)            # 打开fname文件
data.sheet_names()                          # 获取xls文件中所有sheet的名称
table = data.sheet_by_index(0)# 通过索引获取xls文件第0个sheet

sheet=data.get_worksheet_by_name('插入')

table.insert_image('Q5','C:\\Users\\63220\\PycharmProjects\\QQX\\ceshi.png' )



import numpy as np
import matplotlib.pyplot as plt

size = 5
x = np.arange(size)
a = np.random.random(size)
b = np.random.random(size)
c = np.random.random(size)

total_width, n = 0.8, 3
width = total_width / n
x = x - (total_width - width) / 2

plt.bar(x, a,  width=width, label='a')
plt.bar(x + width, b, width=width, label='b')
plt.bar(x + 2 * width, c, width=width, label='c')
plt.legend()
plt.show()


import numpy as np
import matplotlib.pyplot as plt

# a = np.array([5, 20, 15, 25, 10])
# b = np.array([10, 15, 20, 15, 5])

from engine import *
df=pd.read_sql("select zhaoyang,standard,nv_source,nv_source_copy2,nv_standard_copy2 from zhaoyang WHERE zhaoyang is not NULL AND version='{}'".format(now),engine5)
df.dropna(axis=0, how='any', thresh=None, subset=None, inplace=True)
df["differ_day_standard"] = (df["standard"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_source"] = (df["nv_source"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_source_copy2"] = (df["nv_source_copy2"] - df["zhaoyang"]).apply(lambda x: x.days)
df["differ_day_standard_copy2"] = (df["nv_standard_copy2"] - df["zhaoyang"]).apply(lambda x: x.days)
# df["differ_day_source_haomai"] = (df["haomai"] - df["zhaoyang"])
# df["differ_day_source_jinfuzi"] = (df["jinfuzi"] - df["zhaoyang"])

# df.dropna(axis=0, how='any', thresh=None, subset=None, inplace=True)
# axis 指 轴，0是行，1是列，
# how 是删除条件：any 任意一个为na则删除整行/列,all 整行/列为na才删除
# inplace 是否在原DataFrame 上进行删除，false为否
# df.fillna(df.mean())

yy = df[(df.differ_day_source <100 )& (df.differ_day_source>-100)&(df.differ_day_source_copy2<100)&(df.differ_day_source_copy2>-100)&(df.differ_day_standard_copy2<100)&(df.differ_day_standard_copy2>-100)&(df.differ_day_standard<100)&(df.differ_day_standard>-100)]
zz=yy.iloc[:,[5,6,7,8]]
d=zz[(zz.differ_day_standard!=0)|(zz.differ_day_source_copy2!=0)|(zz.differ_day_source!=0)|(zz.differ_day_standard_copy2!=0)]
standard=to_list(d["differ_day_standard"])
source=to_list(d["differ_day_source"])
source_copy2=to_list(d["differ_day_source_copy2"])
standard_copy2=to_list(d["differ_day_standard_copy2"])
zhaoyang=[0]*len(standard)
y=range(len(standard))


x1=sorted(standard)
x2=sorted(source)
x3=sorted(source_copy2)
x4=sorted(standard_copy2)



plt.barh(range(len(x1)), x1)
plt.barh(range(len(x2)), x2)
plt.barh(range(len(x3)), x3)
plt.barh(range(len(x4)), x4)



plt.show()





